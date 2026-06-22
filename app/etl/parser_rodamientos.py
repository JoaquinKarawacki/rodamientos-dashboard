from datetime import datetime, date
from openpyxl import load_workbook


def _normalizar(texto) -> str:
    """
    Colapsa saltos de linea y espacios multiples a un solo espacio.
    'Cat Del\\n(última)' -> 'Cat Del (última)'
    Resuelve el problema de los encabezados con \\n del Excel.
    """
    if texto is None:
        return ""
    return " ".join(str(texto).split())


def _a_fecha(valor) -> date | None:
    """
    Convierte una celda a date. Maneja:
      - datetime / date (openpyxl con celdas de fecha reales)
      - str en formato ISO 'YYYY-MM-DD'
      - str en formato 'DD/MM/YYYY' (algunas celdas de Excel quedan como texto)
    """
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        val = valor.strip()
        try:
            return date.fromisoformat(val[:10])  # YYYY-MM-DD
        except ValueError:
            pass
        try:
            return datetime.strptime(val[:10], "%d/%m/%Y").date()  # DD/MM/YYYY
        except ValueError:
            return None
    return None

def _cargar_filas(ruta_archivo: str, nombre_hoja: str) -> list[tuple]:
    """Lee todas las filas de una hoja como lista de tuplas."""
    wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
    ws = wb[nombre_hoja]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    return filas


def _encontrar_header(filas: list[tuple], marcador: str) -> int:
    """
    Devuelve el indice de la fila cuyo primer valor normalizado coincide
    con el marcador. Mas robusto que hardcodear skiprows: si manana el
    Excel tiene una fila mas o menos arriba, igual encuentra el header.
    """
    for i, fila in enumerate(filas):
        if _normalizar(fila[0]) == marcador:
            return i
    raise ValueError(f"No se encontro la fila de encabezados con '{marcador}'")


def parsear_estado_rodamientos(ruta_archivo: str) -> list[dict]:
    """
    Hoja Estado_Rodamientos = el SEED (estado historico, una fila por turbina).
    Sirve para ambos parques: Peralta tiene columnas extra, pero las columnas
    base que leemos existen en los dos.
    """
    filas = _cargar_filas(ruta_archivo, "Estado_Rodamientos")
    idx = _encontrar_header(filas, "WEC")
    col = {_normalizar(nombre): i for i, nombre in enumerate(filas[idx])}

    registros = []
    for fila in filas[idx + 1:]:
        codigo = _normalizar(fila[col["WEC"]])
        if not codigo or not codigo.upper().startswith(("WEC", "PSP")):
            continue  # fila vacia o no es turbina

        registros.append({
            "codigo_turbina":              codigo,
            "categoria_delantera":         _normalizar(fila[col["Cat Del (última)"]])  or "ND",
            "categoria_trasera":           _normalizar(fila[col["Cat Tras (última)"]]) or "ND",
            "fecha":                       _a_fecha(fila[col["Fecha últ. muestra"]]),
            "tipo_evento":                 _normalizar(fila[col["Tipo últ. muestra"]]) or None,
            "cambio_rodamiento_delantero": _a_fecha(fila[col["Cambio Rod. Frontal"]]),
            "cambio_rodamiento_trasero":   _a_fecha(fila[col["Cambio Rod. Trasero"]]),
            "comentarios":                 None,
        })

    return registros


def parsear_nuevo_control(ruta_archivo: str) -> list[dict]:
    """
    Hoja Nuevo_Control_De_Rodamientos = la CARGA MENSUAL.
    Una fila por evento; las turbinas son columnas marcadas con X.
    Genera una inspeccion por cada turbina marcada.
    """
    filas = _cargar_filas(ruta_archivo, "Nuevo_Control_De_Rodamientos")
    idx = _encontrar_header(filas, "Nueva fecha muestra")
    col = {_normalizar(nombre): i for i, nombre in enumerate(filas[idx])}

    # Columnas cuyo nombre es un codigo de turbina (WEC.. o PSP..)
    columnas_turbina = {
        nombre: i for nombre, i in col.items()
        if nombre.upper().startswith(("WEC", "PSP"))
    }

    registros = []
    for fila in filas[idx + 1:]:
        fecha = _a_fecha(fila[col["Nueva fecha muestra"]])
        if fecha is None:
            continue  # fila vacia, no hay evento

        tipo        = _normalizar(fila[col["Nuevo tipo muestra"]]) or None
        cat_del     = _normalizar(fila[col["Nuevo Cat Del"]])  or "ND"
        cat_tras    = _normalizar(fila[col["Nuevo Cat Tras"]]) or "ND"
        comentarios = _normalizar(fila[col["Comentarios nuevos"]]) or None

        # Aca esta la logica de la X: una inspeccion por turbina marcada
        for codigo_turbina, indice in columnas_turbina.items():
            if _normalizar(fila[indice]).upper() == "X":
                registros.append({
                    "codigo_turbina":              codigo_turbina,
                    "fecha":                       fecha,
                    "tipo_evento":                 tipo,
                    "categoria_delantera":         cat_del,
                    "categoria_trasera":           cat_tras,
                    "cambio_rodamiento_delantero": None,
                    "cambio_rodamiento_trasero":   None,
                    "comentarios":                 comentarios,
                })

    return registros